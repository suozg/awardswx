# ui_utils.py
import wx
from wx import MessageBox, OK, ICON_INFORMATION
import threading
import wx.grid
import csv  
from database_logic import connect_to_database, RankingValues
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
from PIL import ImageEnhance
from config import ALL_COLUMN_LABELS


# клас вікна результатів звіту зі зіставленням
class ComparisonReportFrame(wx.Frame):
    def __init__(self, parent, db_path, key, filters, zvit_fields, zvit_dir, input_data, exel_bmp=None):
        super().__init__(parent, title="Звіт: Порівняння зі списком", size=(950, 680))

        self.db_path = db_path
        self.key = key
        self.filters = filters
        self.zvit_fields = zvit_fields
        self.zvit_dir = zvit_dir
        self.input_data = input_data  # Список словників
        self._exel_bmp = exel_bmp

        self.panel = wx.Panel(self)
        self.grid = wx.grid.Grid(self.panel)
        
        # Индикатор прогресса
        self.progress_gauge = wx.Gauge(self.panel, range=100, style=wx.GA_HORIZONTAL)
        self.progress_gauge.SetValue(0)
        
        # Кнопки збереження та закриття
        self.record_count_label = wx.StaticText(self.panel, label="Обробка даних...")
        
        if self._exel_bmp:
            self.excel_button = wx.BitmapButton(self.panel, bitmap=self._exel_bmp, size=(40, 35))
        else:
            self.excel_button = wx.Button(self.panel, label="Excel", size=(60, 35))
        self.excel_button.Bind(wx.EVT_BUTTON, self.on_export_excel)

        self.csv_button = wx.Button(self.panel, label="CSV", size=(60, 35))
        self.csv_button.Bind(wx.EVT_BUTTON, self.on_export_csv)

        self.close_button = wx.Button(self.panel, label="Закрити", size=(80, 35))
        self.close_button.Bind(wx.EVT_BUTTON, lambda e: self.Close())

        # Компонування
        panel_sizer = wx.BoxSizer(wx.VERTICAL)
        panel_sizer.Add(self.grid, 1, wx.EXPAND | wx.ALL, 5)
        panel_sizer.Add(self.progress_gauge, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 5)

        bottom_sizer = wx.BoxSizer(wx.HORIZONTAL)
        bottom_sizer.Add(self.record_count_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT | wx.RIGHT, 10)
        bottom_sizer.AddStretchSpacer(1)
        bottom_sizer.Add(self.excel_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        bottom_sizer.Add(self.csv_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        bottom_sizer.AddStretchSpacer(1)
        bottom_sizer.Add(self.close_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)

        panel_sizer.Add(bottom_sizer, 0, wx.EXPAND | wx.ALL, 5)
        self.panel.SetSizer(panel_sizer)

        self.result_data = []
        self.columns_names = ["Звання", "ПІБ", "Посада", "РНОКПП", "Знайдені нагороди / подання"]
        
        # Запуск обробки у фоновому потоці
        thread = threading.Thread(target=self.process_in_thread, daemon=True)
        thread.start()


    def process_in_thread(self):
        """Отримує дані з БД, порівнює зі списком у фоновому потоці та оновлює GUI."""
        conn = None
        cursor = None
        try:
            wx.CallAfter(self.progress_gauge.SetValue, 10)
            conn, cursor = connect_to_database(self.key, self.db_path)
            if not cursor:
                wx.CallAfter(wx.MessageBox, "Не вдалося підключитися до бази даних.", "Помилка", wx.OK | wx.ICON_ERROR)
                wx.CallAfter(self.progress_gauge.SetValue, 0)
                return

            wx.CallAfter(self.progress_gauge.SetValue, 20)
            query, params, pre_titlestr = build_query(self.filters, self.zvit_fields)
            raw_db_data = fetch_data(cursor, query, params)
            
            wx.CallAfter(self.SetTitle, f"Порівняльний звіт — {pre_titlestr}")
            wx.CallAfter(self.progress_gauge.SetValue, 40)

            db_rows = []
            if raw_db_data:
                for r in raw_db_data:
                    pib = str(r[0]).strip() if r[0] else ""
                    award_or_pres = str(r[9]).strip() if r[9] else (str(r[5]).strip() if len(r) > 5 and r[5] else "")
                    date_dec = str(r[8]).strip() if len(r) > 8 and r[8] else (str(r[4]).strip() if len(r) > 4 and r[4] else "")
                    inn_db = str(r[15]).strip() if len(r) > 15 and r[15] else "" # РНОКПП з бази даних (індекс 15)
                    
                    award_text = f"{award_or_pres} ({date_dec})" if date_dec else award_or_pres
                    pib_norm = " ".join(pib.upper().split())
                    
                    if pib_norm:
                        db_rows.append({
                            "ПІБ_norm": pib_norm,
                            "Результат_БД": award_text,
                            "РНОКПП_БД": inn_db
                        })

            wx.CallAfter(self.progress_gauge.SetValue, 60)

            # ГРУПУВАННЯ записів з БД
            db_grouped_awards = {}
            db_grouped_inn = {}
            
            for item in db_rows:
                norm = item["ПІБ_norm"]
                res = item["Результат_БД"]
                inn_db = item["РНОКПП_БД"]
                
                if norm not in db_grouped_awards:
                    db_grouped_awards[norm] = []
                if res and res not in db_grouped_awards[norm]:
                    db_grouped_awards[norm].append(res)
                
                # Записуємо РНОКПП обов'язково, якщо він є в базі для цього ПІБ
                if inn_db and not db_grouped_inn.get(norm):
                    db_grouped_inn[norm] = inn_db

            db_grouped_final = {
                k: "; ".join(v) for k, v in db_grouped_awards.items() if v
            }

            wx.CallAfter(self.progress_gauge.SetValue, 80)

            # ГОЛОВНИЙ ЦИКЛ: формуємо результуючий список для ВСІХ записів вхідного файлу
            self.result_data = []
            for row in self.input_data:
                pib_norm = row.get("ПІБ_norm", "")
                
                # 1. Нагороди (якщо немає в базі чи за фільтрами, ставимо "—")
                found_awards = db_grouped_final.get(pib_norm, "—")
                
                # 2. РНОКПП: завжди беремо з бази даних, якщо воно там є для цього ПІБ. 
                # Якщо в базі немає — залишаємо те, що було у вхідному файлі.
                row_inn = db_grouped_inn.get(pib_norm, "")
                if not row_inn:
                    row_inn = row.get("РНОКПП", "").strip()
                
                self.result_data.append({
                    "Звання": row.get("Звання", ""),
                    "ПІБ": row.get("ПІБ", ""),
                    "Посада": row.get("Посада", ""),
                    "РНОКПП": row_inn,  # Підтягується з бази для всіх знайдених людей
                    "Знайдені нагороди / подання": found_awards
                })

            wx.CallAfter(self.progress_gauge.SetValue, 90)
            wx.CallAfter(self.update_gui_after_processing)

        except Exception as e:
            wx.CallAfter(wx.MessageBox, f"Помилка при формуванні звіту:\n{e}", "Помилка", wx.OK | wx.ICON_ERROR)
        finally:
            if cursor: 
                try: cursor.close()
                except Exception: pass
            if conn: 
                try: conn.close()
                except Exception: pass
                

    def update_gui_after_processing(self):
        """Оновлює елементи таблиці та UI у головному потоці."""
        self.grid.CreateGrid(len(self.result_data), len(self.columns_names))
        
        for col_idx, col_name in enumerate(self.columns_names):
            self.grid.SetColLabelValue(col_idx, col_name)

        for row_idx, row in enumerate(self.result_data):
            for col_idx, col_name in enumerate(self.columns_names):
                val = row.get(col_name, "")
                self.grid.SetCellValue(row_idx, col_idx, str(val))

        self.grid.AutoSizeColumns()
        self.record_count_label.SetLabel(f"Всього у списку: {len(self.result_data)}")
        self.progress_gauge.SetValue(100)

    def on_export_excel(self, event):
        if not self.result_data:
            return
        
        with wx.FileDialog(self, "Зберегти звіт Excel", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT,
                           defaultDir=self.zvit_dir or "", defaultFile="звіт_порівняння.xlsx") as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                path = dlg.GetPath()
                wb = Workbook()
                ws = wb.active
                ws.title = "Звіт"
                
                ws.append(self.columns_names)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                
                for row_dict in self.result_data:
                    row_vals = [row_dict.get(col, "") for col in self.columns_names]
                    ws.append(row_vals)
                    
                wb.save(path)
                wx.MessageBox(f"Успішно збережено в {path}", "Успіх", wx.OK | wx.ICON_INFORMATION)

    def on_export_csv(self, event):
        if not self.result_data:
            return
            
        with wx.FileDialog(self, "Зберегти звіт CSV", wildcard="CSV files (*.csv)|*.csv",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT,
                           defaultDir=self.zvit_dir or "", defaultFile="звіт_порівняння.csv") as dlg:
            if dlg.ShowModal() == wx.ID_OK:
                path = dlg.GetPath()
                with open(path, mode='w', newline='', encoding='utf-8-sig') as csv_file:
                    writer = csv.writer(csv_file, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                    writer.writerow(self.columns_names)
                    for row_dict in self.result_data:
                        writer.writerow([row_dict.get(col, "") for col in self.columns_names])

                wx.MessageBox(f"Успішно збережено в {path}", "Успіх", wx.OK | wx.ICON_INFORMATION)


def parse_input_person_csv(filepath):
    """
    Завантажує та нормалізує CSV файл зі списком людей (1 людина - 1 рядок).
    Повертає список словників з ключами: ['Звання', 'ПІБ', 'Посада', 'РНОКПП', 'ПІБ_norm', 'РНОКПП_norm']
    """
    spis_rows = []
    with open(filepath, "r", encoding="utf-8-sig", errors="ignore") as f:
        sample = f.read(2048)
        f.seek(0)
        delimiter = ';' if ';' in sample else ','
        reader = csv.reader(f, delimiter=delimiter)
        
        for row in reader:
            if not row or len(row) < 1:
                continue
            
            pib = ""
            inn = ""
            zvannya = ""
            posada = ""
            
            for cell in row:
                cell_clean = cell.strip()
                if not cell_clean:
                    continue
                if len(cell_clean) == 10 and cell_clean.isdigit():
                    inn = cell_clean
                elif len(cell_clean.split()) >= 2 and not pib:
                    pib = cell_clean
                elif not zvannya:
                    zvannya = cell_clean
                else:
                    posada = cell_clean

            if pib:
                pib_norm = " ".join(pib.strip().upper().split())
                inn_norm = inn.strip() if inn.strip().isdigit() else ""
                
                spis_rows.append({
                    "Звання": zvannya,
                    "ПІБ": pib,
                    "Посада": posada,
                    "РНОКПП": inn,
                    "ПІБ_norm": pib_norm,
                    "РНОКПП_norm": inn_norm
                })

    return spis_rows


# Пошук і відображення списку нагород з автопідстановкою
class ComboSearchHelper:
    def __init__(self, combo_ctrl, debounce_delay=600, kartka_panel_instance=None):
        self.combo_ctrl = combo_ctrl
        self.DEBOUNCE_DELAY_MS = debounce_delay
        self._loaded_items = []
        self._is_user_typing = False
        self.search_timer = None
        self.kartka_panel = kartka_panel_instance 

        self.combo_ctrl.Bind(wx.EVT_TEXT, self.on_text_changed)
        self.combo_ctrl.Bind(wx.EVT_COMBOBOX, self.on_combo_selected)

    def set_items(self, names):
        self._loaded_items = names

    def on_text_changed(self, event):
        if self.kartka_panel and self.kartka_panel._is_programmatic_award_change:
            event.Skip()
            return

        self._is_user_typing = True
        if self.search_timer and self.search_timer.IsRunning():
            self.search_timer.Stop()
        self.search_timer = wx.CallLater(self.DEBOUNCE_DELAY_MS, self._perform_search_and_update)
        event.Skip()

    def on_combo_selected(self, event):
        if self.search_timer and self.search_timer.IsRunning():
            self.search_timer.Stop()
        self._is_user_typing = False
        event.Skip()

    def _perform_search_and_update(self):
        current_text = self.combo_ctrl.GetValue()
        if not current_text:
            filtered = self._loaded_items
        else:
            filtered = [a for a in self._loaded_items 
                       if a.lower().startswith(current_text.lower()) or 
                          current_text.lower() in a.lower()]

        self.combo_ctrl.Unbind(wx.EVT_TEXT, handler=self.on_text_changed)

        insertion_point = self.combo_ctrl.GetInsertionPoint()

        self.combo_ctrl.SetItems(filtered)
        self.combo_ctrl.SetValue(current_text)
        self.combo_ctrl.SetInsertionPoint(insertion_point)

        is_exact_match_selected = current_text in filtered and len(filtered) == 1

        if self._is_user_typing and current_text and filtered and not is_exact_match_selected:
            try:
                self.combo_ctrl.Popup()
            except Exception:
                pass
        else:
            try:
                self.combo_ctrl.Dismiss()
            except Exception:
                pass

        self.combo_ctrl.Bind(wx.EVT_TEXT, self.on_text_changed)
        self._is_user_typing = False

    def reset_search(self):
        self._is_user_typing = False
        if self.search_timer and self.search_timer.IsRunning():
            self.search_timer.Stop()

        self.combo_ctrl.Unbind(wx.EVT_TEXT, handler=self.on_text_changed)
        self.combo_ctrl.ChangeValue("")
        self.combo_ctrl.SetItems(self._loaded_items)
        self.combo_ctrl.SetSelection(0)
        self.combo_ctrl.Bind(wx.EVT_TEXT, self.on_text_changed)


# ----------------- ЛОГІКА ЗАПИТІВ ДЛЯ ЗВІТІВ

MIN_ROW_HEIGHT = 25
MAX_WIDTH = 300 

class ReportGeneratorWx(wx.Frame):
    def __init__(self, parent, db_path, key, zvit_fields, zvit_dir, filters, exel_bmp=None):
        super().__init__(parent, title="Звіт", size=(800, 680))

        self.ALL_COLUMN_LABELS = ALL_COLUMN_LABELS

        self.filters = filters
        self.db_path = db_path
        self.key = key
        self._exel_bmp = exel_bmp

        if isinstance(zvit_fields, str):
            zvit_fields = list(map(int, zvit_fields.split(',')))
        self.zvit_fields = zvit_fields
        self.zvit_dir = zvit_dir
        self.data = None

        self.panel = wx.Panel(self)
        self.rows_expanded = False 
        self.active_columns = []

        self.wrap_renderer = wx.grid.GridCellAutoWrapStringRenderer()

        num_standard_flags = 14

        for i in range(num_standard_flags):
            original_data_index = i
            if self.zvit_fields[i] == 1:
                self.active_columns.append(original_data_index)

        last_flag_index_in_zvit = 14
        last_flag = self.zvit_fields[last_flag_index_in_zvit]

        original_index_rno = 15
        original_index_dob = 16

        if last_flag == 1:
            if original_index_rno < len(self.ALL_COLUMN_LABELS):
                 self.active_columns.append(original_index_rno)

        elif last_flag == 2:
             if original_index_dob < len(self.ALL_COLUMN_LABELS):
                  self.active_columns.append(original_index_dob)

        elif last_flag == 3:
             if original_index_rno < len(self.ALL_COLUMN_LABELS):
                  self.active_columns.append(original_index_rno)
             if original_index_dob < len(self.ALL_COLUMN_LABELS):
                  self.active_columns.append(original_index_dob)

        try:
             self.column_labels = [self.ALL_COLUMN_LABELS[i] for i in self.active_columns]
        except IndexError:
             self.column_labels = ["Помилка в конфігурації колонок"]
             self.active_columns = [0]

        self.grid = wx.grid.Grid(self.panel)
        grid_col_count = len(self.active_columns) if self.active_columns else 1
        self.grid.CreateGrid(0, grid_col_count)

        for idx, label in enumerate(self.column_labels):
            self.grid.SetColLabelValue(idx, label)

        self.grid.EnableEditing(False)

        self.grid.Bind(wx.grid.EVT_GRID_CELL_RIGHT_CLICK, self.on_grid_right_click)
        self.grid.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, self.on_grid_cell_dclick)

        # Индикатор прогресса
        self.progress_gauge = wx.Gauge(self.panel, range=100, style=wx.GA_HORIZONTAL)
        self.progress_gauge.SetValue(0)

        self.record_count_label = wx.StaticText(self.panel)

        self.excel_button = wx.BitmapButton(self.panel, bitmap=self._exel_bmp, size=(40, 35))
        if self._exel_bmp and self._exel_bmp.IsOk():
            self.excel_button.SetBitmap(self._exel_bmp, wx.LEFT)
            self.excel_button.SetLabel("")
        self.excel_button.Bind(wx.EVT_BUTTON, self.on_export_excel)

        self.csv_button = wx.Button(self.panel, label="CSV", size=(50, 35))
        self.csv_button.Bind(wx.EVT_BUTTON, self.on_export_csv)

        self.close_button = wx.Button(self.panel, label="Закрити")
        self.close_button.Bind(wx.EVT_BUTTON, self.on_close)

        self._last_selected_blocks = None
        self._last_selected_cells = None
        self._last_selected_rows = None
        self._last_right_clicked_cell = None

        panel_sizer = wx.BoxSizer(wx.VERTICAL)
        panel_sizer.Add(self.grid, 1, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 5)
        panel_sizer.Add(self.progress_gauge, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 5)

        bottom_sizer = wx.BoxSizer(wx.HORIZONTAL)
        bottom_sizer.Add(self.record_count_label, 0, wx.ALIGN_CENTER_VERTICAL | wx.LEFT | wx.RIGHT, 5)
        bottom_sizer.AddStretchSpacer(1)
        bottom_sizer.Add(self.excel_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        bottom_sizer.Add(self.csv_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        bottom_sizer.AddStretchSpacer(1)       
        bottom_sizer.Add(self.close_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        panel_sizer.Add(bottom_sizer, 0, wx.EXPAND | wx.ALL, 5)

        self.panel.SetSizer(panel_sizer)
        self.Layout()

        self._load_data()

    def on_close(self, event):
        self.Close()

    def on_grid_cell_dclick(self, event):
        event.Skip(False)

    def on_grid_right_click(self, event):
        try:
            self._last_selected_blocks = list(self.grid.GetSelectedBlocks())
        except TypeError:
            self._last_selected_blocks = []

        self._last_selected_cells = self.grid.GetSelectedCells()
        self._last_selected_rows = self.grid.GetSelectedRows()
        self._last_right_clicked_cell = (event.GetRow(), event.GetCol())

        menu = wx.Menu()
        copy_cells_id = wx.NewIdRef()
        copy_rows_id = wx.NewIdRef()

        copy_cells_item = menu.Append(copy_cells_id, "Копіювати виділені ячейки")
        if len(self._last_selected_blocks) > 0 or len(self._last_selected_cells) > 0:
             self.Bind(wx.EVT_MENU, self.on_copy_cells, id=copy_cells_id)
        else:
             copy_cells_item.Enable(False)

        copy_rows_item = menu.Append(copy_rows_id, "Копіювати виділені рядки")
        if len(self._last_selected_rows) > 0:
             self.Bind(wx.EVT_MENU, self.on_copy_rows, id=copy_rows_id)
        else:
             copy_rows_item.Enable(False)

        if copy_cells_item.IsEnabled() or copy_rows_item.IsEnabled() or (self._last_right_clicked_cell[0] >= 0 and self._last_right_clicked_cell[1] >= 0):
             if menu.GetMenuItemCount() > 0:
                 menu.AppendSeparator()

        toggle_rows_id = wx.NewIdRef()
        toggle_label = "Згорнути рядки" if self.rows_expanded else "Розгорнути рядки"
        menu.Append(toggle_rows_id, toggle_label)
        self.Bind(wx.EVT_MENU, self.on_toggle_row_height, id=toggle_rows_id)

        self.grid.PopupMenu(menu, event.GetPosition())

    def on_copy_cells(self, event):
        selected_blocks = self._last_selected_blocks if self._last_selected_blocks is not None else []
        selected_cells = self._last_selected_cells if self._last_selected_cells is not None else []

        if len(selected_blocks) == 0 and len(selected_cells) == 0:
            wx.MessageBox("Немає виділених ячеек для копіювання.", "Інформація", wx.OK | wx.ICON_INFORMATION)
            return

        text_to_copy = ""
        temp_data = {}

        if len(selected_blocks) > 0:
             for block in selected_blocks:
                 r_start = block.GetTopLeft().GetRow()
                 c_start = block.GetTopLeft().GetCol()
                 r_end = block.GetBottomRight().GetRow()
                 c_end = block.GetBottomRight().GetCol()
                 for r in range(r_start, r_end + 1):
                     for c in range(c_start, c_end + 1):
                         temp_data[(r, c)] = self.grid.GetCellValue(r, c)

        if selected_cells:
            for r, c in selected_cells:
                temp_data[(r, c)] = self.grid.GetCellValue(r, c)

        if temp_data:
            all_coords = list(temp_data.keys())
            min_row = min(r for r, c in all_coords)
            max_row = max(r for r, c in all_coords)
            min_col = min(c for r, c in all_coords)
            max_col = max(c for r, c in all_coords)
            lines = []
            for r in range(min_row, max_row + 1):
                row_values = []
                for c in range(min_col, max_col + 1):
                    value = temp_data.get((r, c), "")
                    row_values.append(value)
                lines.append('\t'.join(row_values))
            text_to_copy = '\n'.join(lines)
        else:
            text_to_copy = ""

        if text_to_copy:
           if wx.TheClipboard.Open():
               clipboard_data = wx.TextDataObject(text_to_copy)
               wx.TheClipboard.SetData(clipboard_data)
               wx.TheClipboard.Close()
           else:
               wx.MessageBox("Не вдалося отримати доступ до буфера обміну.", "Помилка", wx.OK | wx.ICON_ERROR)

    def on_copy_rows(self, event):
        selected_rows = self._last_selected_rows if self._last_selected_rows is not None else []

        if not selected_rows:
            wx.MessageBox("Немає виділених рядків для копіювання.", "Інформація", wx.OK | wx.ICON_INFORMATION)
            return

        lines = []
        for row_idx in sorted(selected_rows):
             if row_idx < 0 or row_idx >= self.grid.GetNumberRows():
                 continue

             row_values = []
             for col_idx in range(self.grid.GetNumberCols()):
                  value = self.grid.GetCellValue(row_idx, col_idx)
                  row_values.append(str(value) if value is not None else "")

             lines.append('\t'.join(row_values))

        text_to_copy = '\n'.join(lines)

        if text_to_copy:
           if wx.TheClipboard.Open():
               clipboard_data = wx.TextDataObject(text_to_copy)
               wx.TheClipboard.SetData(clipboard_data)
               wx.TheClipboard.Close()
           else:
               wx.MessageBox("Не вдалося отримати доступ до буфера обміну.", "Помилка", wx.OK | wx.ICON_ERROR)

    def on_copy_clicked_cell(self, event):
        if self._last_right_clicked_cell and self._last_right_clicked_cell[0] >= 0 and self._last_right_clicked_cell[1] >= 0:
            r, c = self._last_right_clicked_cell
            try:
                text_to_copy = self.grid.GetCellValue(r, c)
                if text_to_copy or text_to_copy == "":
                    if wx.TheClipboard.Open():
                        clipboard_data = wx.TextDataObject(text_to_copy)
                        wx.TheClipboard.SetData(clipboard_data)
                        wx.TheClipboard.Close()
                    else:
                        wx.MessageBox("Не вдалося отримати доступ до буфера обміну.", "Помилка", wx.OK | wx.ICON_ERROR)

            except Exception as e:
                wx.MessageBox(f"Помилка при копіюванні ячейки: {e}", "Помилка", wx.OK | wx.ICON_ERROR)
        else:
            wx.MessageBox("Не вдалося визначити ячейку для копіювання.", "Помилка", wx.OK | wx.ICON_ERROR)

    def on_toggle_row_height(self, event):
        if self.grid is None or self.grid.GetNumberRows() == 0:
            return

        if self.rows_expanded:
            self.rows_expanded = False
            min_height = self.grid.GetDefaultRowSize() if self.grid.GetDefaultRowSize() > 0 else MIN_ROW_HEIGHT
            for row_idx in range(self.grid.GetNumberRows()):
                self.grid.SetRowSize(row_idx, min_height)
        else:
            self.rows_expanded = True
            self.grid.AutoSizeRows()

        self.grid.ForceRefresh()
        self.panel.Layout()

    def _update_grid_with_data(self):
        if self.data is None or not self.data:
            self.grid.ClearGrid()
            if self.grid.GetNumberRows() > 0:
                self.grid.DeleteRows(0, self.grid.GetNumberRows())
            self.record_count_label.SetLabel("Немає даних для відображення")
            self.SetTitle("Звіт – Немає даних")
            self.grid.ForceRefresh()
            self.panel.Layout()
            return

        self.SetTitle(f"Звіт – {self.pre_titlestr}")

        num_rows_needed = len(self.data)
        num_cols_needed = len(self.active_columns)

        self.grid.ClearGrid()
        current_rows = self.grid.GetNumberRows()
        if current_rows < num_rows_needed:
             self.grid.AppendRows(num_rows_needed - current_rows)
        elif current_rows > num_rows_needed:
             self.grid.DeleteRows(num_rows_needed, current_rows - num_rows_needed)

        current_cols = self.grid.GetNumberCols()
        if current_cols < num_cols_needed:
             self.grid.AppendCols(num_cols_needed - current_cols)
        elif current_cols > num_cols_needed:
             self.grid.DeleteCols(num_cols_needed, current_cols - num_cols_needed)

        if self.grid.GetNumberCols() == len(self.column_labels):
             for idx, label in enumerate(self.column_labels):
                 self.grid.SetColLabelValue(idx, label)

        for row_idx, row in enumerate(self.data):
            max_needed_data_index = max(self.active_columns) if self.active_columns else -1

            if len(row) > max_needed_data_index:
                 for col_idx, data_index in enumerate(self.active_columns):
                    try:
                        value = row[data_index]
                        self.grid.SetCellValue(row_idx, col_idx, str(value if value is not None else ''))
                    except IndexError:
                         self.grid.SetCellValue(row_idx, col_idx, "ПОМИЛКА ДАНИХ (IndexError)")
                    except Exception:
                         self.grid.SetCellValue(row_idx, col_idx, "ПОМИЛКА ДАНИХ")
            else:
                for col_idx, data_index in enumerate(self.active_columns):
                    if data_index < len(row):
                        try:
                            value = row[data_index]
                            self.grid.SetCellValue(row_idx, col_idx, str(value if value is not None else ''))
                        except Exception:
                            self.grid.SetCellValue(row_idx, col_idx, "ПОМИЛКА ДАНИХ")

        original_data_index_to_wrap = 5

        if original_data_index_to_wrap in self.active_columns:
            col_idx_visible = self.active_columns.index(original_data_index_to_wrap)

            wrap_attr = wx.grid.GridCellAttr()
            wrap_attr.SetRenderer(self.wrap_renderer)
            wrap_attr.SetAlignment(wx.ALIGN_LEFT, wx.ALIGN_TOP)

            self.grid.SetColSize(col_idx_visible, MAX_WIDTH)
            self.grid.SetColAttr(col_idx_visible, wrap_attr)
               
        if self.rows_expanded:
             self.grid.AutoSizeRows()
        else:
             min_height = self.grid.GetDefaultRowSize() if self.grid.GetDefaultRowSize() > 0 else MIN_ROW_HEIGHT
             for row_idx in range(self.grid.GetNumberRows()):
                self.grid.SetRowSize(row_idx, min_height)

        self.record_count_label.SetLabel(f"Кількість записів: {len(self.data)}")
        self.grid.ForceRefresh()
        self.panel.Layout()

    def _load_data(self):
        thread = threading.Thread(target=self._load_data_background, daemon=True)
        thread.start()
        self.record_count_label.SetLabel("Чекайте, йде обробка запиту...")
        wx.CallAfter(self.record_count_label.SetLabel, "Чекайте, йде обробка запиту...")

    def _load_data_background(self):
        conn = None
        cursor = None
        try:
            wx.CallAfter(self.progress_gauge.SetValue, 10)
            conn, cursor = connect_to_database(self.key, self.db_path)
            if conn is None or cursor is None:
                wx.CallAfter(self._show_message, "Не вдалося підключитися до бази даних.")
                wx.CallAfter(self.record_count_label.SetLabel, "Помилка підключення до БД")
                wx.CallAfter(self.progress_gauge.SetValue, 0)
                wx.CallAfter(self._update_grid_with_data)
                return

            wx.CallAfter(self.progress_gauge.SetValue, 40)
            query, params, pre_titlestr = build_query(self.filters, self.zvit_fields)
            self.pre_titlestr = pre_titlestr

            if not query:
                 wx.CallAfter(self._show_message, "Не вдалося сформувати запит до бази даних.")
                 wx.CallAfter(self.record_count_label.SetLabel, "Помилка формування запиту")
                 wx.CallAfter(self.progress_gauge.SetValue, 0)
                 wx.CallAfter(self._update_grid_with_data)
                 return

            wx.CallAfter(self.progress_gauge.SetValue, 70)
            data = fetch_data(cursor, query, params)
            self.data = data

            wx.CallAfter(self.progress_gauge.SetValue, 100)
            wx.CallAfter(self._update_grid_with_data)

            if not data:
                wx.CallAfter(self._show_message, "Немає даних, що відповідають критеріям фільтрації.")

        except Exception as e:
            error_message = f"Помилка під час виконання запиту або обробки даних: {e}"
            wx.CallAfter(self._show_message, error_message)
            wx.CallAfter(self.record_count_label.SetLabel, "Помилка під час обробки даних")
            wx.CallAfter(self.progress_gauge.SetValue, 0)
            wx.CallAfter(self._update_grid_with_data)

        finally:
            if cursor:
                try: cursor.close()
                except Exception as e: print(f"Error closing cursor: {e}")
            if conn:
                try: conn.close()
                except Exception as e: print(f"Error closing connection: {e}")

    def _show_message(self, message):
        if hasattr(self, 'IsBeingDeleted') and self.IsBeingDeleted():
            return
        if self and not self.IsBeingDeleted():
            wx.MessageBox(message, "Інформація", wx.OK | wx.ICON_INFORMATION)

    def on_export_excel(self, event):
        if self.data is None or not self.data:
            wx.MessageBox("Немає даних для експорту.", "Інформація", wx.OK | wx.ICON_INFORMATION)
            return

        defaultDir = self.zvit_dir if self.zvit_dir else os.getcwd()
        with wx.FileDialog(self, "Зберегти звіт у Excel",
                           wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT,
                           defaultDir=defaultDir,
                           defaultFile="звіт.xlsx") as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            pathname = fileDialog.GetPath()
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Звіт"

                if hasattr(self, 'pre_titlestr') and self.pre_titlestr:
                    ws.append([self.pre_titlestr])
                    ws['A1'].font = Font(bold=True)
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.column_labels))
                    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    header_row_index = 2
                else:
                    header_row_index = 1

                ws.append(self.column_labels)
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for col_idx, cell in enumerate(ws[header_row_index]):
                    cell.font = header_font
                    cell.alignment = header_alignment
                    col_letter = get_column_letter(col_idx + 1)
                    ws.column_dimensions[col_letter].width = 20

                data_alignment = Alignment(vertical='top', wrap_text=True)
                for row in self.data:
                    excel_row_data = []
                    for data_index in self.active_columns:
                        try:
                            value = row[data_index] if data_index < len(row) else ""
                            excel_row_data.append(str(value if value is not None else ''))
                        except Exception:
                            excel_row_data.append("Error")

                    ws.append(excel_row_data)

                PIXELS_TO_EXCEL_UNITS = 7
                MAX_EXCEL_WIDTH = MAX_WIDTH / PIXELS_TO_EXCEL_UNITS
                self.long_text_source_indices = [5] 

                for i, original_col_index in enumerate(self.active_columns):
                    col_letter = get_column_letter(i + 1)
                    if original_col_index in self.long_text_source_indices:
                        ws.column_dimensions[col_letter].width = MAX_EXCEL_WIDTH

                for row_idx in range(header_row_index + 1, ws.max_row + 1):
                     for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).alignment = data_alignment

                wb.save(pathname)
                wx.CallAfter(wx.MessageBox, f"Звіт успішно збережено в {pathname}", "Успіх", wx.OK | wx.ICON_INFORMATION)

            except Exception as e:
                error_msg = f"Помилка при збереженні звіту: {e}"
                wx.CallAfter(wx.MessageBox, error_msg, "Помилка", wx.OK | wx.ICON_ERROR)

    def on_export_csv(self, event):
        if self.data is None or not self.data:
            wx.MessageBox("Немає даних для експорту.", "Інформація", wx.OK | wx.ICON_INFORMATION)
            return

        defaultDir = self.zvit_dir if self.zvit_dir else os.getcwd()
        with wx.FileDialog(self, "Зберегти звіт у CSV",
                           wildcard="CSV files (*.csv)|*.csv",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT,
                           defaultDir=defaultDir,
                           defaultFile="звіт.csv") as fileDialog:

            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return

            pathname = fileDialog.GetPath()
            try:
                with open(pathname, mode='w', newline='', encoding='utf-8-sig') as csv_file:
                    writer = csv.writer(csv_file, delimiter=';', quoting=csv.QUOTE_MINIMAL)

                    if hasattr(self, 'pre_titlestr') and self.pre_titlestr:
                        writer.writerow([self.pre_titlestr])
                        writer.writerow([])

                    writer.writerow(self.column_labels)

                    for row in self.data:
                        csv_row_data = []
                        for data_index in self.active_columns:
                            try:
                                value = row[data_index] if data_index < len(row) else ""
                                csv_row_data.append(str(value if value is not None else ''))
                            except Exception:
                                csv_row_data.append("Error")

                        writer.writerow(csv_row_data)

                wx.CallAfter(wx.MessageBox, f"Звіт успішно збережено в {pathname}", "Успіх", wx.OK | wx.ICON_INFORMATION)

            except Exception as e:
                error_msg = f"Помилка при збереженні CSV файлу: {e}"
                wx.CallAfter(wx.MessageBox, error_msg, "Помилка", wx.OK | wx.ICON_ERROR)


def build_query(filters, zvit_fields):
    params = {}
    not_protokol = ""

    mode = filters.get('mode')
    specific_submission = filters.get('specific_submission') if mode == 'submission' else False

    if specific_submission:
        params = {}
        submission_number_full = filters.get('submission_number', '')
        number_pres = submission_number_full.split(' від')[0] if ' від' in submission_number_full else submission_number_full

        params["number_pres"] = number_pres
        pre_titlestr = f"ПОДАННЯ №{filters.get('submission_number', '')}"

        id_order_name_alph = "ORDER BY CASE WHEN p.name GLOB '[А-Д]*' THEN 1 WHEN p.name GLOB 'Є*' THEN 2 WHEN p.name GLOB '[Е-З]*' THEN 3 WHEN p.name GLOB 'І*' THEN 4 WHEN p.name GLOB '[И-Я]*' THEN 5 ELSE 6 END, p.name"

        query = f"""
            SELECT p.name, p.rank, p.unit, pr.registration,
                 pr.date_registration, pr.text_presentation, pr.report,
                 m.decree, m.date_decree,
                 a.denotation,
                 m.date_handover,
                 p2.name,
                 m.consignment_note, m.number_meed, m.dead, p.inn, p.date_birth
            FROM presentation pr
            JOIN personality p ON pr.id_personality = p.id
            LEFT JOIN meed m ON pr.id_meed = m.id
            LEFT JOIN award a ON m.id_award = a.id
            LEFT JOIN personality p2 ON p2.id = m.handover
            WHERE pr.registration = :number_pres
            {id_order_name_alph}
            """
        return query, params, pre_titlestr

    start_date = filters.get("start_date")
    end_date = filters.get("end_date")
    params.update({"date_0": start_date, "date_1": end_date})

    id_order_name_alph = "ORDER BY \
                 CASE \
                   WHEN p.name GLOB '[А-Д]*' THEN 1 \
                   WHEN p.name GLOB 'Є*' THEN 2 \
                   WHEN p.name GLOB '[Е-З]*' THEN 3 \
                   WHEN p.name GLOB 'І*' THEN 4 \
                   WHEN p.name GLOB '[И-Я]*' THEN 5 \
                   ELSE 6 \
                 END, p.name"

    unit_sel = filters.get("unit", "")
    unit_condition = ""
    if unit_sel:
        unit_condition = "AND p.unit = :unit"
        params["unit"] = unit_sel

    id_rank_condition = ""
    rank_sel = ""
    RankRegularStr = []
    rank_params = {}

    person_category = filters.get("person_category")
    civilian = filters.get("civilian")

    if person_category == 1:
        RankRegularStr = ['%полковник', 'майор', 'капітан%', '%лейтенант']
        rank_sel = "офіцери"
    elif person_category == 2:
        RankRegularStr = ['%сержант']
        rank_sel = "сержанти"
    elif person_category == 3:
        RankRegularStr = ['%солдат', '%матрос']
        rank_sel = "солдати"
    elif person_category == 4:
        id_rank_condition = "AND LENGTH(IFNULL(p.rank , '')) <> 0"
        rank_sel = "усі військові"
    elif civilian:
        id_rank_condition = "AND (p.rank IS NULL OR TRIM(p.rank) = '')"
        rank_sel = "цивільні"
        unit_sel = "" 
        unit_condition = "" 
    else:
        rank_sel = "усі особи"
        unit_sel = "" 
        unit_condition = "" 

    if RankRegularStr:
        conditions = []
        for i, pattern in enumerate(RankRegularStr):
            key = f"rank{i}"
            conditions.append(f"LOWER(p.rank) LIKE :{key}")
            rank_params[key] = pattern.lower()
        id_rank_condition = "AND (" + " OR ".join(conditions) + ")"

    params.update(rank_params)

    if mode == "awarding":
        handover_status = filters.get('handover_status')
        awarding_where_conditions = []
        award_params_awarding = {}

        awarding_dead_status_condition = ""
        if filters.get('posthumous'): 
            awarding_dead_status_condition = "m.dead = '1'" 

        presentation_subquery_where = "" 

        if not filters.get("all_time") and handover_status != 1:
            awarding_where_conditions.append("m.date_decree BETWEEN :date_0 AND :date_1")

        if handover_status == 3:
            pre_titlestr = f'ПРИЗНАЧЕНІ нагороди {unit_sel}'
        else:
            awarding_where_conditions.append("m.consignment_note IS NOT NULL AND m.consignment_note <> ''")
            consignment_note_filter = filters.get("consignment_note")

            if handover_status == 1:
                awarding_where_conditions.append("(m.handover IS NULL OR m.handover = '')")
                pre_titlestr = f"ЗАЛИШОК нагород В НАТУРІ {unit_sel}"

                if consignment_note_filter:
                     awarding_where_conditions.append("m.consignment_note = :consignment_note")
                     award_params_awarding["consignment_note"] = consignment_note_filter

            elif handover_status == 2:
                 awarding_where_conditions.append("m.handover IS NOT NULL AND m.handover <> ''")
                 pre_titlestr = f'ВРУЧЕНІ нагороди {unit_sel} {not_protokol}'

                 if consignment_note_filter:
                      awarding_where_conditions.append("m.consignment_note = :consignment_note")
                      award_params_awarding["consignment_note"] = consignment_note_filter

        if unit_condition:
            awarding_where_conditions.append(unit_condition)

        if id_rank_condition:
            awarding_where_conditions.append(id_rank_condition)

        id_rank_award_condition = ""
        is_issue_protocols_filter = filters.get("issue_protocols", False)
        target_ranks_for_protocol_indices = []

        if is_issue_protocols_filter:
            try:
                target_ranks_for_protocol_indices.append(RankingValues.index("Найвища"))
                target_ranks_for_protocol_indices.append(RankingValues.index("Державні"))
                target_ranks_for_protocol_indices.append(RankingValues.index("Відомчі"))
            except ValueError:
                target_ranks_for_protocol_indices = []
        
        if is_issue_protocols_filter and target_ranks_for_protocol_indices:
            id_rank_award_condition = (
                f"a.ranking IN ({','.join([':' + f'ranking_p{i}' for i in range(len(target_ranks_for_protocol_indices))])})"
                f" AND (m.handover IS NULL OR m.handover NOT LIKE '%$%')"
            )
            for i, rank_idx in enumerate(target_ranks_for_protocol_indices):
                award_params_awarding[f"ranking_p{i}"] = rank_idx
            not_protokol = "(відсутні протоколи видачі)"
        else:
            award_rank_filter_value = filters.get("award_rank")
            if award_rank_filter_value and award_rank_filter_value.strip():
                try:
                    award_ranking_value = RankingValues.index(award_rank_filter_value)
                    id_rank_award_condition = "a.ranking = :ranking"
                    award_params_awarding["ranking"] = award_ranking_value
                except ValueError:
                    pass

        if id_rank_award_condition:
            awarding_where_conditions.append(id_rank_award_condition)        

        if awarding_dead_status_condition:
            awarding_where_conditions.append(awarding_dead_status_condition)

        award_id_filter_value = filters.get("award_id")
        if award_id_filter_value is not None:
            awarding_where_conditions.append("m.id_award = :award_id")
            award_params_awarding["award_id"] = award_id_filter_value

        awarding_where_string = ""
        if awarding_where_conditions:
            cleaned_conditions = [cond.replace("AND ", "", 1) if cond.startswith("AND ") else cond for cond in awarding_where_conditions]
            awarding_where_string = "WHERE " + " AND ".join(cleaned_conditions)

        params.update(award_params_awarding)

        # query = f"""
        #     SELECT
        #         p.name, p.rank, p.unit,
        #         pr.registration, pr.date_registration, pr.text_presentation, pr.report,
        #         m.decree, m.date_decree,
        #         a.denotation,
        #         m.date_handover,
        #         p2.name,
        #         m.consignment_note, m.number_meed, m.dead, p.inn, p.date_birth
        #     FROM meed m
        #     JOIN personality p ON m.id_personality = p.id
        #     JOIN award a ON m.id_award = a.id
        #     LEFT JOIN (
        #         SELECT id, id_personality, id_meed, worker,
        #           registration, date_registration, text_presentation, report
        #         FROM presentation
        #         {presentation_subquery_where}
        #     ) pr ON m.id = pr.id_meed
        #     LEFT JOIN personality p2 ON p2.id = m.handover
        #     {awarding_where_string}
        #     {id_order_name_alph}
        #     """
        query = f"""
            SELECT
                p.name, p.rank, p.unit,
                pr.registration, pr.date_registration, pr.text_presentation, pr.report,
                m.decree, m.date_decree,
                a.denotation,
                m.date_handover,
                p2.name,
                m.consignment_note, m.number_meed, m.dead, p.inn, p.date_birth
            FROM personality p
            LEFT JOIN meed m ON m.id_personality = p.id
            LEFT JOIN award a ON m.id_award = a.id
            LEFT JOIN (
                SELECT id, id_personality, id_meed, worker,
                  registration, date_registration, text_presentation, report
                FROM presentation
                {presentation_subquery_where}
            ) pr ON m.id = pr.id_meed
            LEFT JOIN personality p2 ON p2.id = m.handover
            {awarding_where_string}
            {id_order_name_alph}
            """
    
    elif mode == 'submission':
        id_worker_condition = ""
        id_worker_map = {"Усі": "_", "ВП": "0", "МПЗ": "1", "Інші": "2"}
        worker_filter_value = filters.get("worker", "")
        id_worker_value = id_worker_map.get(worker_filter_value, "")

        worker_filter_active = worker_filter_value and id_worker_value != "_" 
        if worker_filter_active:
            id_worker_condition = "AND pr.worker = :worker_id"
            params["worker_id"] = int(id_worker_value)

        id_meed_status = ""
        label_selected_presvarR2 = "УСІ подання"

        submission_status_filter = filters.get('submission_status')

        if submission_status_filter == 1:
            id_meed_status = "AND pr.id_meed > 0"
            label_selected_presvarR2 = "ПОГОДЖЕНІ подання"
        elif submission_status_filter == 2:
            id_meed_status = "AND pr.id_meed = '0'"
            label_selected_presvarR2 = "НЕ ПОГОДЖЕНІ подання"
        elif submission_status_filter == 3:
            id_meed_status = "AND (pr.id_meed IS NULL OR TRIM(pr.id_meed) = '')"
            label_selected_presvarR2 = "подання У ЗАЛИШКУ"

        pre_titlestr = f'{label_selected_presvarR2} -вик.{filters.get("worker", "")}'

        if filters.get('posthumous'):
            id_dead_status = "AND pr.report == 'посмертно'"  
            pre_titlestr += " -посмертно"
        else:
           id_dead_status = ""

        pre_titlestr += f" -{filters.get('start_date')}_{filters.get('end_date')}"

        if filters.get('person_category'):
            pre_titlestr += f" -{filters.get('person_category')}"

        if filters.get('unit'):
            pre_titlestr += f" -{filters.get('unit')}"        

        submission_conditions_list = []

        if not filters.get("all_time"):
            submission_conditions_list.append("pr.date_registration BETWEEN :date_0 AND :date_1")

        if id_meed_status:
            cleaned_meed_status = id_meed_status.lstrip().replace("AND ", "", 1).strip()
            if cleaned_meed_status:
                submission_conditions_list.append(cleaned_meed_status)

        if id_dead_status:
            cleaned_dead_status = id_dead_status.lstrip().replace("AND ", "", 1).strip()
            if cleaned_dead_status:
                submission_conditions_list.append(cleaned_dead_status)

        if unit_condition:
             cleaned_unit_condition = unit_condition.lstrip().replace("AND ", "", 1).strip()
             if cleaned_unit_condition:
                 submission_conditions_list.append(cleaned_unit_condition)

        if id_rank_condition:
             cleaned_rank_condition = id_rank_condition.lstrip().replace("AND ", "", 1).strip()
             if cleaned_rank_condition:
                 submission_conditions_list.append(cleaned_rank_condition)

        if id_worker_condition:
             cleaned_worker_condition = id_worker_condition.lstrip().replace("AND ", "", 1).strip()
             if cleaned_worker_condition:
                 submission_conditions_list.append(cleaned_worker_condition)

        submission_where_string = ""
        if submission_conditions_list:
            submission_where_string = "WHERE " + " AND ".join(submission_conditions_list)

        # query = f"""
        #      SELECT p.name, p.rank, p.unit, pr.registration,
        #           pr.date_registration, pr.text_presentation, pr.report,
        #           m.decree, m.date_decree,
        #           a.denotation,
        #           m.date_handover,
        #           CASE WHEN p2.name IS NOT NULL THEN p2.name ELSE m.handover END,
        #           m.consignment_note, m.number_meed, m.dead, p.inn, p.date_birth
        #      FROM presentation pr
        #      JOIN personality p ON pr.id_personality = p.id
        #      LEFT JOIN meed m ON m.id = pr.id_meed
        #      LEFT JOIN award a ON m.id_award = a.id
        #      LEFT JOIN personality p2 ON p2.id = m.handover
        #      {submission_where_string}
        #      {id_order_name_alph}
        #      """
        query = f"""
             SELECT p.name, p.rank, p.unit, pr.registration,
                  pr.date_registration, pr.text_presentation, pr.report,
                  m.decree, m.date_decree,
                  a.denotation,
                  m.date_handover,
                  CASE WHEN p2.name IS NOT NULL THEN p2.name ELSE m.handover END,
                  m.consignment_note, m.number_meed, m.dead, p.inn, p.date_birth
             FROM personality p
             LEFT JOIN presentation pr ON pr.id_personality = p.id
             LEFT JOIN meed m ON m.id = pr.id_meed
             LEFT JOIN award a ON m.id_award = a.id
             LEFT JOIN personality p2 ON p2.id = m.handover
             {submission_where_string}
             {id_order_name_alph}
             """    

    else:
        query = ""
        params = {}
        pre_titlestr = "Не вибрано режим вибірки"

    return query, params, pre_titlestr

def fetch_data(cursor, query, params=None):
    try:
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        result = cursor.fetchall()
        return result

    except Exception:
        return None


def load_image_from_blob(image_blob, max_dim=80, grayscale=False, brightness_factor=1.0):
    if image_blob is None:
        return wx.Bitmap(max_dim, max_dim)

    try:
        img_stream = io.BytesIO(image_blob)
        image = Image.open(img_stream)
        width, height = image.size

        if max(width, height) > max_dim:
            scale = max_dim / max(width, height)
            new_width = int(width * scale)
            new_height = int(height * scale)
            try:
                image = image.resize((new_width, new_height), resample=Image.Resampling.LANCZOS)
            except AttributeError:
                image = image.resize((new_width, new_height), resample=Image.LANCZOS)
        elif max(width, height) == 0:
            return wx.Bitmap(max_dim, max_dim)

        image = image.convert("RGBA")

        if grayscale:
            image = image.convert("LA")

        if brightness_factor != 1.0:
            if image.mode == "LA":
                image = image.convert("RGBA")
            enhancer = ImageEnhance.Brightness(image)
            image = enhancer.enhance(brightness_factor)

        if image.mode == "RGBA":
            rgb_data = image.convert("RGB").tobytes()
            alpha_data = image.getchannel("A").tobytes()
            wx_image = wx.Image(image.width, image.height)
            wx_image.SetData(rgb_data)
            wx_image.SetAlpha(alpha_data)
        elif image.mode == "LA":
            image = image.convert("RGBA")
            rgb_data = image.convert("RGB").tobytes()
            alpha_data = image.getchannel("A").tobytes()
            wx_image = wx.Image(image.width, image.height)
            wx_image.SetData(rgb_data)
            wx_image.SetAlpha(alpha_data)
        elif image.mode == "L":
            rgb_data = image.convert("RGB").tobytes()
            wx_image = wx.Image(image.width, image.height)
            wx_image.SetData(rgb_data)
        else:
            rgb_data = image.convert("RGB").tobytes()
            wx_image = wx.Image(image.width, image.height)
            wx_image.SetData(rgb_data)

        bitmap = wx.Bitmap(wx_image)
        return bitmap

    except Exception:
        error_bitmap = wx.Bitmap(max_dim, max_dim)
        dc = wx.MemoryDC(error_bitmap)
        dc.SetBackground(wx.Brush(wx.RED))
        dc.Clear()
        dc.SetTextForeground(wx.WHITE)
        dc.DrawText("Помилка", 5, 5)
        dc.SelectObject(wx.NullBitmap)
        return error_bitmap


def on_highlight(richtext_ctrl, word_to_find, text_to_highlight, highlight_color):
    start_index = 0
    text_attr = wx.TextAttr()
    text_attr.SetTextColour(highlight_color)

    while True:
        start_index = text_to_highlight.find(word_to_find, start_index)
        if start_index == wx.NOT_FOUND:
            break

        end_index = start_index + len(word_to_find)
        richtext_ctrl.SetStyle(start_index, end_index, text_attr)
        start_index = end_index
